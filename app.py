"""
TCL Appraisal Model
"""

import copy
import sys
from pathlib import Path
from io import BytesIO

import pandas as pd
import streamlit as st
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

# --- Make engine/ importable without changing any engine files ---
ENGINE_DIR = Path(__file__).resolve().parent / "engine"
sys.path.insert(0, str(ENGINE_DIR))

from load_assumptions import load_assumptions
from inputs import (
    development_mix as default_development_mix,
    place_scenario_options,
    place_scenario_controls as default_place_scenario_controls,
    place_scenario_user_inputs as default_place_scenario_user_inputs,
    crime_inputs as default_crime_inputs,
    land_infra_options,
    land_infra_inputs as default_land_infra_inputs,
    commercial_floorspace_inputs as default_commercial_floorspace_inputs, pbsa_inputs as default_pbsa_inputs,
    get_floorspace,
)
from additionality import additionality_questions as default_additionality_questions
from calculations import (
    population_by_typology, demographic_employment_outputs, get_total_gia,
    transport_emissions_mode_calculations, transport_emissions_valuation,
    embodied_carbon, environmental_quality, health_wellbeing, civic_engagement,
    travel_time_and_costs, cost_of_crime, economic_activity,
    construction_activity, fiscal, land_values,
    public_infrastructure, commercial_floorspace_indicator, pbsa_indicator,
)
from results import build_dashboard
from guidance import (
    GUIDANCE_TITLE, GUIDANCE_INTRO, INPUT_SECTIONS, GUIDANCE_NOTE,
    GUIDANCE_RESULTS_INTRO, RESULTS_SECTIONS,
)
from notes import (
    PLACE_SCENARIO_NOTES, PLACE_SCENARIO_USER_INPUT_NOTES,
    CRIME_NOTES, LAND_INFRA_NOTES, ADDITIONALITY_NOTES,
)
from branding import inject_brand_css, render_brand_table, brand_callout, render_header, render_stat_row, force_light_mode
from admin import render_admin_panel

st.set_page_config(page_title="TCL Appraisal Model", layout="wide")
inject_brand_css()
force_light_mode()


# =================================================================================================
# HELPERS
# =================================================================================================

def gbp(x):
    if isinstance(x, str):
        return x
    return f"£{x:,.0f}"

def gbp_compact(x):
    if isinstance(x, str):
        return x
    abs_x = abs(x)
    if abs_x >= 1_000_000:
        return f"£{x/1_000_000:,.1f}m"
    elif abs_x >= 1_000:
        return f"£{x/1_000:,.0f}k"
    else:
        return f"£{x:,.0f}"

def pct(x):
    if isinstance(x, str):
        return x
    return f"{x:.0%}"


@st.cache_resource
def get_assumptions():
    return load_assumptions()


def optional_number(label, dict_ref, key_name, widget_key, min_value=0.0, step=1.0, fmt="%.0f", show_caption=True, help=None):
    current = dict_ref[key_name]
    provided = st.checkbox(f"Provide: {label}", value=current is not None, key=f"{widget_key}_toggle")
    if provided:
        value = st.number_input(label, value=float(current) if current is not None else 0.0,
                                 min_value=min_value, step=step, format=fmt, key=widget_key, help=help)
        dict_ref[key_name] = value
    else:
        if show_caption:
            st.caption(f"{label}: not provided, assumptions default will be used where applicable")
        dict_ref[key_name] = None


def render_guidance_sections(sections: list[dict], body_key: str) -> None:
    """Renders a list of guidance sections. body_key is 'what_you_need' or 'description'."""
    for section in sections:
        st.subheader(section["title"])
        body = section[body_key]
        if isinstance(body, list):
            for item in body:
                st.markdown(f"- {item}")
        else:
            st.markdown(body)

        if "extra" in section:
            st.markdown(f"**{section['extra']['heading']}**")
            for item in section["extra"]["items"]:
                st.markdown(f"- {item}")

        if section.get("why"):
            st.caption(section["why"])
        st.divider()


# =================================================================================================
# EXCEL EXPORT FORMATTING
# =================================================================================================

CURRENCY_FORMAT = '"£"#,##0'
PERCENT_FORMAT = '0.0%'

PERCENT_LABELS = {"Deadweight", "Displacement", "Leakage", "Net factor"}

MONEY_LABELS = {
    "Gross £ value", "Net additional £ value",
    "Gross annual wellbeing value", "Net additional annual wellbeing value",
    "Gross annual health value", "Net additional annual health value",
    "Gross annual value", "Net additional annual value",
    "Gross TC spend", "Net TC spend", "Gross GVA", "Net GVA", "Net additional GVA",
    "Gross fiscal value", "Net additional fiscal value",
    "Gross land value uplift", "Net additional land value uplift",
    "Gross infrastructure savings (one-off)", "Net infrastructure savings (one-off)",
    "Gross annual revenue saving", "Net annual revenue saving",
    "Gross total GVA", "Net total GVA", "Net total spend",
}

PERCENT_COLUMNS = {"Baseline share", "TCL share"}
MONEY_COLUMNS = {"Gross value", "Net value", "Gross GVA"}


def _append_row(ws, values, formats=None):
    ws.append(values)
    row = ws[ws.max_row]
    if formats:
        for cell, fmt in zip(row, formats):
            if fmt and isinstance(cell.value, (int, float)):
                cell.number_format = fmt


def _format_measure_value_sheet(ws):
    for row in ws.iter_rows(min_row=2):
        label_cell, value_cell = row[0], row[1]
        if not isinstance(value_cell.value, (int, float)):
            continue
        if label_cell.value in PERCENT_LABELS:
            value_cell.number_format = PERCENT_FORMAT
        elif label_cell.value in MONEY_LABELS:
            value_cell.number_format = CURRENCY_FORMAT


def _format_columns_by_header(ws, percent_columns=None, money_columns=None):
    percent_columns = percent_columns or set()
    money_columns = money_columns or set()
    if ws.max_row < 2:
        return
    headers = [c.value for c in ws[1]]
    for col_idx, header in enumerate(headers):
        fmt = PERCENT_FORMAT if header in percent_columns else CURRENCY_FORMAT if header in money_columns else None
        if fmt:
            for row in ws.iter_rows(min_row=2, min_col=col_idx + 1, max_col=col_idx + 1):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = fmt


def build_excel_export(dashboard: dict, sheets: dict[str, pd.DataFrame]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Impact Dashboard"

    ws.append(["Net Additional Social Value"])
    _append_row(ws, ["Annual", dashboard["social_value"]["Annual"]], [None, CURRENCY_FORMAT])
    _append_row(ws, ["10-Year NPV", dashboard["social_value"]["10-Year NPV"]], [None, CURRENCY_FORMAT])
    _append_row(ws, ["Social value per home (10yr NPV)", dashboard["social_value"]["Social value per home (10yr NPV)"]],
                [None, CURRENCY_FORMAT])
    ws.append([])

    ws.append(["Social Value Sensitivity (+/-10%)"])
    ws.append(["Scenario", "Annual social value", "10yr Social value NPV"])
    for scenario, values in dashboard["social_value_sensitivity"].items():
        _append_row(ws, [scenario, values["Annual social value"], values["10yr Social value NPV"]],
                    [None, CURRENCY_FORMAT, CURRENCY_FORMAT])
    ws.append([])

    ws.append(["Social Value Composition: Annual Impacts"])
    ws.append(["Impact Area", "Annual", "% of annual total"])
    for area, values in dashboard["composition_annual"].items():
        _append_row(ws, [area, values["Annual"], values["% of annual total"]],
                    [None, CURRENCY_FORMAT, PERCENT_FORMAT])
    ws.append([])

    ws.append(["Social Value Composition: One-off Impacts"])
    for area, value in dashboard["composition_oneoff"].items():
        _append_row(ws, [area, value], [None, CURRENCY_FORMAT])
    ws.append([])

    ws.append(["Core Development Outputs"])
    for k, v in dashboard["core_outputs"].items():
        ws.append([k, v])
    ws.append([])

    ws.append(["Net Additional Economic and Fiscal Value"])
    for measure, values in dashboard["econ_fiscal"].items():
        if isinstance(values, dict):
            annual_val = values.get("Annual", "-")
            npv_val = values.get("10-Year NPV", "-")
            _append_row(ws, [measure, annual_val, npv_val],
                        [None, CURRENCY_FORMAT, CURRENCY_FORMAT])
        else:
            _append_row(ws, [measure, "-", values], [None, None, CURRENCY_FORMAT])

    ws.append(["Economic and Fiscal Value Sensitivity (+/-10%)"])
    ws.append(["Measure", "-10%", "Central", "+10%"])
    for measure, values in dashboard["econ_fiscal_sensitivity"].items():
        _append_row(ws, [measure, values["-10%"], values["Central"], values["+10%"]],
                    [None, CURRENCY_FORMAT, CURRENCY_FORMAT, CURRENCY_FORMAT])
    ws.append([])

    ws.append(["One-off Construction Impacts"])
    con_o = dashboard["construction_oneoff"]
    _append_row(ws, ["Net additional GVA", con_o["Net additional GVA"]], [None, CURRENCY_FORMAT])
    ws.append(["Net additional PYE jobs", con_o["Net additional PYE jobs"]])
    ws.append([])

    ws.append(["Commercial Floorspace Impacts Sensitivity (+/-10%)"])
    ws.append(["Measure", "-10%", "Central", "+10%"])
    for measure, values in dashboard["commercial_floorspace_sensitivity"].items():
        _append_row(ws, [measure, values["-10%"], values["Central"], values["+10%"]],
                    [None, CURRENCY_FORMAT, CURRENCY_FORMAT, CURRENCY_FORMAT])

    for name, df in sheets.items():
        sheet = wb.create_sheet(name[:31])
        for row in dataframe_to_rows(df, index=True, header=True):
            sheet.append(row)
        if df.index.name == "Measure":
            _format_measure_value_sheet(sheet)
        else:
            _format_columns_by_header(sheet, percent_columns=PERCENT_COLUMNS, money_columns=MONEY_COLUMNS)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =================================================================================================
# SESSION STATE INIT
# =================================================================================================

if "development_mix" not in st.session_state:
    st.session_state.development_mix = copy.deepcopy(default_development_mix)
if "place_scenario_controls" not in st.session_state:
    st.session_state.place_scenario_controls = copy.deepcopy(default_place_scenario_controls)
if "place_scenario_user_inputs" not in st.session_state:
    st.session_state.place_scenario_user_inputs = copy.deepcopy(default_place_scenario_user_inputs)
if "crime_inputs" not in st.session_state:
    st.session_state.crime_inputs = copy.deepcopy(default_crime_inputs)
if "land_infra_inputs" not in st.session_state:
    st.session_state.land_infra_inputs = copy.deepcopy(default_land_infra_inputs)
if "commercial_floorspace_inputs" not in st.session_state:
    st.session_state.commercial_floorspace_inputs = copy.deepcopy(default_commercial_floorspace_inputs)
if "pbsa_inputs" not in st.session_state:
    st.session_state.pbsa_inputs = copy.deepcopy(default_pbsa_inputs)
if "additionality_questions" not in st.session_state:
    st.session_state.additionality_questions = copy.deepcopy(default_additionality_questions)
if "show_results" not in st.session_state:
    st.session_state.show_results = False
if "guidance_done" not in st.session_state:
    st.session_state.guidance_done = False

assumptions = get_assumptions()
render_admin_panel(get_assumptions.clear)

development_mix = st.session_state.development_mix
place_scenario_controls = st.session_state.place_scenario_controls
place_scenario_user_inputs = st.session_state.place_scenario_user_inputs
crime_inputs = st.session_state.crime_inputs
land_infra_inputs = st.session_state.land_infra_inputs
commercial_floorspace_inputs = st.session_state.commercial_floorspace_inputs
pbsa_inputs = st.session_state.pbsa_inputs
additionality_questions = st.session_state.additionality_questions


# =================================================================================================
# TITLE
# =================================================================================================

render_header("Town Centre Living Appraisal Model")


# =================================================================================================
# GUIDANCE PAGE
# =================================================================================================

if not st.session_state.guidance_done:
    st.header(GUIDANCE_TITLE)
    st.markdown(GUIDANCE_INTRO)

    render_guidance_sections(INPUT_SECTIONS, body_key="what_you_need")

    brand_callout(GUIDANCE_NOTE)

    if st.button("Start", type="primary", use_container_width=True):
        st.session_state.guidance_done = True
        st.rerun()

# =================================================================================================
# INPUTS — ONE PAGE, TABBED
# =================================================================================================

elif not st.session_state.show_results:
    input_tabs = st.tabs([
        "Development Mix", "Place & Scenario", "Crime", "Land & Infrastructure",
        "Commercial Floorspace", "Additionality Questions",
    ])

    # --- Development Mix ---
    with input_tabs[0]:
        st.caption("Enter the number of homes per typology. Floor area are optional, leave at 0 to use the assumptions default.")
        for typology in development_mix:
            st.subheader(typology)
            c1, c2, c3, c4 = st.columns(4)
            development_mix[typology]["Private Homes"] = c1.number_input(
                "Private Homes", min_value=0, value=int(development_mix[typology]["Private Homes"]),
                key=f"dm_{typology}_ph")
            development_mix[typology]["Social/Affordable Homes"] = c2.number_input(
                "Social/Affordable Homes", min_value=0, value=int(development_mix[typology]["Social/Affordable Homes"]),
                key=f"dm_{typology}_sh")
            development_mix[typology]["Private Floor Area per unit (m2)"] = c3.number_input(
                "Private floor area (m2)", min_value=0.0,
                value=float(development_mix[typology]["Private Floor Area per unit (m2)"]),
                key=f"dm_{typology}_pf")
            development_mix[typology]["Social Floor Area per unit (m2)"] = c4.number_input(
                "Social floor area (m2)", min_value=0.0,
                value=float(development_mix[typology]["Social Floor Area per unit (m2)"]),
                key=f"dm_{typology}_sf")

        st.divider()
        gia = get_total_gia(development_mix, assumptions)
        population = population_by_typology(development_mix, assumptions)
        c1, c2, c3 = st.columns(3)
        c1.metric("Total private GIA", f"{gia['Total private GIA']:,.0f} m²")
        c2.metric("Total social GIA", f"{gia['Total social GIA']:,.0f} m²")
        c3.metric("Total residents", f"{population['Total']['Total residents']:,.0f}")

    # --- Place & Scenario ---
    with input_tabs[1]:
        for question, options in place_scenario_options.items():
            current = place_scenario_controls[question]
            idx = options.index(current) if current in options else 0
            place_scenario_controls[question] = st.selectbox(
                question, options, index=idx, key=f"psc_{question}",
                help=PLACE_SCENARIO_NOTES.get(question))

        st.divider()
        st.subheader("Numerical inputs")

        place_scenario_user_inputs["How many people live within 250m - 500m of the proposed development"] = st.number_input(
            "How many people live within 250m - 500m of the proposed development",
            min_value=0, value=int(place_scenario_user_inputs["How many people live within 250m - 500m of the proposed development"]),
            key="psui_pop",
            help=PLACE_SCENARIO_USER_INPUT_NOTES.get("How many people live within 250m - 500m of the proposed development"))

        place_scenario_user_inputs["Estimate the capital costs of the town centre living development"] = st.number_input(
            "Estimate the capital costs of the town centre living development (£)",
            min_value=0, value=int(place_scenario_user_inputs["Estimate the capital costs of the town centre living development"]),
            key="psui_capex",
            help=PLACE_SCENARIO_USER_INPUT_NOTES.get("Estimate the capital costs of the town centre living development"))

        car_key = "% of households in new TCL development with 1+ car"
        current_override = place_scenario_user_inputs[car_key]
        use_override = st.checkbox("Override % of households with 1+ car (otherwise uses assumptions default)",
                                    value=current_override is not None, key="psui_car_toggle")
        if use_override:
            val = st.slider("% of households with 1+ car", 0.0, 1.0,
                             value=float(current_override) if current_override is not None else 0.5,
                             step=0.01, key="psui_car_val")
            place_scenario_user_inputs[car_key] = val
        else:
            place_scenario_user_inputs[car_key] = None
            st.caption("Will be estimated from the development mix and assumptions car ownership rates.")

    # --- Crime ---
    with input_tabs[2]:
        optional_number("Town Centre Population", crime_inputs, "Town Centre Population", "crime_pop",
                         show_caption=False, help=CRIME_NOTES.get("Town Centre Population"))
        st.divider()
        for category in ["Violence", "Theft", "Criminal damage", "ASB"]:
            st.subheader(category)
            optional_number(f"{category} Recorded Incidents", crime_inputs[category], "Recorded Incidents", f"crime_{category}",
                             show_caption=False,
                             help=CRIME_NOTES.get(category, CRIME_NOTES.get("Recorded Incidents")))

    # --- Land & Infrastructure ---
    with input_tabs[3]:
        q = "Are you able to provide details on the GDV and existing land value"
        options = land_infra_options[q]
        current = land_infra_inputs[q]
        idx = options.index(current) if current in options else 1
        land_infra_inputs[q] = st.selectbox(q, options, index=idx, key="lii_qualifier",
                                             help=LAND_INFRA_NOTES.get(q))

        if land_infra_inputs[q].strip() == "Yes":
            optional_number("Total Gross Development Value (GDV) £", land_infra_inputs,
                             "Total Gross Development Value (GDV) £", "lii_gdv", step=1000.0, show_caption=False,
                             help=LAND_INFRA_NOTES.get("Total Gross Development Value (GDV) £"))
            optional_number("Existing use / baseline land value (£)", land_infra_inputs,
                             "Existing use / baseline land value (£)", "lii_baseline", step=1000.0, show_caption=False,
                             help=LAND_INFRA_NOTES.get("Existing use / baseline land value (£)"))
        else:
            st.caption("Land Values indicator will show 'N/A' since GDV details aren't being provided.")

    # --- Commercial Floorspace ---
    with input_tabs[4]:
        st.caption("Optional")
        for category in commercial_floorspace_inputs:
            commercial_floorspace_inputs[category] = st.number_input(
                f"{category} (m²)", min_value=0.0, value=float(commercial_floorspace_inputs[category]),
                key=f"cf_{category}")

        st.divider()
        st.subheader("Purpose Built Student Accommodation (PBSA)")
        st.caption("Optional. Enter the number of rooms only — do not include floorspace above for PBSA.")
        pbsa_key = "Purpose Built Student Accommodation - Number of Rooms"
        pbsa_inputs[pbsa_key] = st.number_input(
            "Number of rooms", min_value=0, value=int(pbsa_inputs[pbsa_key]), key="pbsa_rooms")

    
    # --- Additionality Questions ---
    with input_tabs[5]:
        for impact_area, factors in additionality_questions.items():
            st.subheader(impact_area)
            for factor, details in factors.items():
                options = details["options"]
                current = details["answer"]
                idx = options.index(current) if current in options else 0
                details["answer"] = st.selectbox(
                    f"{factor}: {details['question']}", options, index=idx,
                    key=f"aq_{impact_area}_{factor}",
                    help=ADDITIONALITY_NOTES.get(f"{impact_area}_{factor}"))
            st.divider()

        st.divider()
        if st.button("Show Results", type="primary", use_container_width=True):
            st.session_state.show_results = True
            st.rerun()


# =================================================================================================
# RESULTS
# =================================================================================================

if st.session_state.guidance_done and st.session_state.show_results:

    if st.button("Edit Inputs", use_container_width=True):
        st.session_state.show_results = False
        st.rerun()

    with st.expander("How to read these results", expanded=False):
        st.markdown(GUIDANCE_RESULTS_INTRO)
        render_guidance_sections(RESULTS_SECTIONS, body_key="description")

    try:
        dashboard = build_dashboard(
            development_mix, place_scenario_controls, place_scenario_user_inputs,
            crime_inputs, land_infra_inputs, commercial_floorspace_inputs,
            additionality_questions, assumptions,
        )
    except (TypeError, ZeroDivisionError, KeyError):
        st.warning(
            "Not enough information has been entered to calculate results. "
            "Please go back to Edit Inputs and check every section has been completed, "
            "especially any numeric fields."
        )
        st.stop()

    sheets = {} 

    st.header("Impact Dashboard")

    st.subheader("Net Additional Social Value")
    c1, c2, c3 = st.columns(3)
    c1.metric("Annual", gbp_compact(dashboard["social_value"]["Annual"]))
    c2.metric("10-Year NPV", gbp_compact(dashboard["social_value"]["10-Year NPV"]))
    c3.metric("Social value per home (10yr NPV)", gbp_compact(dashboard["social_value"]["Social value per home (10yr NPV)"]))

    st.markdown("**Social Value Sensitivity (±10%)**")
    sens_rows = []
    for scenario, values in dashboard["social_value_sensitivity"].items():
        sens_rows.append({"Scenario": scenario, "Annual social value": gbp_compact(values["Annual social value"]),
                           "10yr Social value NPV": gbp_compact(values["10yr Social value NPV"])})
    render_brand_table(pd.DataFrame(sens_rows))

    st.markdown("**Social Value Composition: Annual Impacts**")
    comp_rows = []
    for area, values in dashboard["composition_annual"].items():
        comp_rows.append({"Impact Area": area, "Annual": gbp_compact(values["Annual"]),
                           "% of annual total": pct(values["% of annual total"])})
    render_brand_table(pd.DataFrame(comp_rows))

    st.markdown("**Social Value Composition: One-off Impacts**")
    render_stat_row([(area, gbp_compact(value)) for area, value in dashboard["composition_oneoff"].items()])

    st.subheader("Core Development Outputs")
    co = dashboard["core_outputs"]
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Total Units", f"{co['Total Units']:,.0f}")
    c2.metric("Total Residents", f"{co['Total Residents']:,.0f}")
    c3.metric("Working-age Adults", f"{co['Working-age Adults']:,.0f}")
    c4.metric("Employed Adults", f"{co['Employed Adults']:,.0f}")
    c5.metric("FTE jobs", f"{co['FTE jobs']:,.0f}")

    st.subheader("Net Additional Economic and Fiscal Value")
    ef_rows = []
    for measure, values in dashboard["econ_fiscal"].items():
        if isinstance(values, dict):
            annual_val = gbp_compact(values["Annual"]) if "Annual" in values else "-"
            npv_val = gbp_compact(values["10-Year NPV"]) if "10-Year NPV" in values else "-"
            ef_rows.append({"Measure": measure, "Annual": annual_val, "10-Year NPV": npv_val})
        else:
            ef_rows.append({"Measure": measure, "Annual": "-", "10-Year NPV": gbp_compact(values)})

    st.markdown("**Economic and Fiscal Value Sensitivity (±10%)**")
    efs_rows = []
    for measure, values in dashboard["econ_fiscal_sensitivity"].items():
        efs_rows.append({"Measure": measure, "-10%": gbp_compact(values["-10%"]), "Central": gbp_compact(values["Central"]), "+10%": gbp_compact(values["+10%"])})
    render_brand_table(pd.DataFrame(efs_rows))

    st.subheader("One-off Construction Impacts")
    con_o = dashboard["construction_oneoff"]
    c1, c2 = st.columns(2)
    c1.metric("Net additional GVA", gbp_compact(con_o["Net additional GVA"]))
    c2.metric("Net additional PYE jobs", f"{con_o['Net additional PYE jobs']:,.0f}")

    st.subheader("Commercial Floorspace Impacts")
    
    cfs_rows = []
    for measure, values in dashboard["commercial_floorspace_sensitivity"].items():
        if "FTE jobs" in measure: 
            cfs_rows.append({"Measure": measure, "-10%": f"{values['-10%']:,.0f}", "Central": f"{values['Central']:,.0f}", "+10%": f"{values['+10%']:,.0f}"})
        else:    
            cfs_rows.append({"Measure": measure, "-10%": gbp_compact(values["-10%"]), "Central": gbp_compact(values["Central"]), "+10%": gbp_compact(values["+10%"])})
    render_brand_table(pd.DataFrame(cfs_rows))

    st.subheader("PBSA Impacts (Supplementary)")
    pbsa_dashboard = pbsa_indicator(pbsa_inputs, additionality_questions, assumptions)
    c1, c2 = st.columns(2)
    c1.metric("Net total FTE jobs", f"{pbsa_dashboard['Net total FTE jobs']:,.0f}")
    c2.metric("Net total GVA", gbp_compact(pbsa_dashboard["Net total GVA"]))

    st.divider()
    st.header("Detailed Impacts")

    with st.expander("1. Transport Emissions"):
        mode_calcs = transport_emissions_mode_calculations(place_scenario_controls, development_mix, assumptions)
        df = pd.DataFrame(mode_calcs["modes"]).T
        df_display = df.rename(columns={"Baseline share": "Baseline %", "TCL share": "TCL %"})
        df_display["Baseline %"] = df_display["Baseline %"].apply(lambda x: f"{x:.0%}")
        df_display["TCL %"] = df_display["TCL %"].apply(lambda x: f"{x:.0%}")
        for c in ["Baseline km", "TCL km", "Emission factor", "Baseline tCO2e", "TCL tCO2e", "Savings tCO2e"]:
            df_display[c] = df_display[c].apply(lambda x: f"{x:,.0f}")
        render_brand_table(df_display.reset_index().rename(columns={"index": "Mode"}))
        st.metric("Gross savings (tCO2e)", f"{mode_calcs['Gross savings']:,.0f}")

        v = transport_emissions_valuation(place_scenario_controls, development_mix, additionality_questions, assumptions)
        sheets["1. Transport Emissions"] = pd.DataFrame(
            list({"Gross tCO2e saved": v["Gross tCO2e saved"], "Net additional tCO2e saved": v["Net additional tCO2e saved"],
                  "Gross £ value": v["Gross £ value"], "Net additional £ value": v["Net additional £ value"]}.items()),
            columns=["Measure", "Value"]).set_index("Measure")
        c1, c2, c3 = st.columns(3)
        c1.metric("Gross £ value", gbp_compact(v["Gross £ value"]))
        c2.metric("Net factor", pct(v["Net factor"]))
        c3.metric("Net additional £ value", gbp_compact(v["Net additional £ value"]))
        st.write(f"Deadweight {pct(v['Deadweight'])} - Displacement {pct(v['Displacement'])} - Leakage {pct(v['Leakage'])}")

    with st.expander("2. Embodied Carbon"):
        ec = embodied_carbon(development_mix, place_scenario_controls, additionality_questions, assumptions)
        sheets["2. Embodied Carbon"] = pd.DataFrame(
            list({"Total GIA": ec["Total GIA"], "Gross tCO2e saved": ec["Gross tCO2e saved"],
                  "Net additional tCO2e saved": ec["Net additional tCO2e saved"],
                  "Gross £ value": ec["Gross £ value"], "Net additional £ value": ec["Net additional £ value"]}.items()),
            columns=["Measure", "Value"]).set_index("Measure")
        c1, c2, c3 = st.columns(3)
        c1.metric("Total private GIA", f"{ec['Total private GIA']:,.0f} m²")
        c2.metric("Total social GIA", f"{ec['Total social GIA']:,.0f} m²")
        c3.metric("Total GIA", f"{ec['Total GIA']:,.0f} m²")
        c1, c2 = st.columns(2)
        c1.metric("Proposed carbon factor", f"{ec['Proposed carbon factor']:,.0f}")
        c2.metric("Comparator (baseline) factor", f"{ec['Comparator carbon factor']:,.0f}")
        c1, c2, c3 = st.columns(3)
        c1.metric("Gross tCO2e saved", f"{ec['Gross tCO2e saved']:,.0f}")
        c2.metric("Gross £ value", gbp_compact(ec["Gross £ value"]))
        c3.metric("Net additional £ value", gbp_compact(ec["Net additional £ value"]))
        st.write(f"Deadweight {pct(ec['Deadweight'])} - Displacement {pct(ec['Displacement'])} - Leakage {pct(ec['Leakage'])}")

    with st.expander("3. Environmental Quality"):
        eq = environmental_quality(place_scenario_controls, place_scenario_user_inputs, additionality_questions, assumptions)
        sheets["3. Environmental Quality"] = pd.DataFrame(
            list({"Gross annual wellbeing value": eq["Gross annual wellbeing value"],
                  "Net additional annual wellbeing value": eq["Net additional annual wellbeing value"]}.items()),
            columns=["Measure", "Value"]).set_index("Measure")
        c1, c2, c3 = st.columns(3)
        c1.metric("Population within 500m", f"{eq['Population within 500m']:,.0f}")
        c2.metric("Brownfield / gap site impact", f"{eq['Brownfield / gap site impact']:.4f}")
        c3.metric("Vacant units impact", f"{eq['Vacant units impact']:.4f}")
        c1, c2 = st.columns(2)
        c1.metric("Gross annual wellbeing value", gbp_compact(eq["Gross annual wellbeing value"]))
        c2.metric("Net additional annual value", gbp_compact(eq["Net additional annual wellbeing value"]))
        st.write(f"Deadweight {pct(eq['Deadweight'])} - Displacement {pct(eq['Displacement'])} - Leakage {pct(eq['Leakage'])}")

    with st.expander("4. Resident Population"):
        population = population_by_typology(development_mix, assumptions)
        df = pd.DataFrame(population).T
        sheets["4. Resident Population"] = df.copy()
        df_display = df.copy()
        for c in df_display.columns:
            df_display[c] = df_display[c].apply(lambda x: f"{x:,.0f}")
        render_brand_table(df_display.reset_index().rename(columns={"index": "Typology"}))
        demographics = demographic_employment_outputs(development_mix, assumptions)
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Children", f"{demographics['Children']:,.0f}")
        c2.metric("Working-age adults", f"{demographics['Working-age adults']:,.0f}")
        c3.metric("Older adults", f"{demographics['Older adults']:,.0f}")
        c4.metric("Employed adults", f"{demographics['Employed adults']:,.0f}")

    with st.expander("5. Health & Wellbeing"):
        hw = health_wellbeing(place_scenario_controls, development_mix, additionality_questions, assumptions)
        sheets["5. Health and Wellbeing"] = pd.DataFrame(
            list({"Gross annual health value": hw["Gross annual health value"],
                  "Net additional annual health value": hw["Net additional annual health value"]}.items()),
            columns=["Measure", "Value"]).set_index("Measure")
        c1, c2 = st.columns(2)
        c1.metric("Total residents (weighted)", f"{hw['Total residents (weighted)']:,.0f}")
        c2.metric("Minutes uplift / person / week", f"{hw['Minutes uplift per person/week']:,.0f}")
        c1, c2 = st.columns(2)
        c1.metric("Residents newly meeting threshold", f"{hw['Residents newly meeting threshold']:,.0f}")
        c2.metric("Residents already active", f"{hw['Residents already active']:,.0f}")
        c1, c2 = st.columns(2)
        c1.metric("Gross annual health value", gbp_compact(hw["Gross annual health value"]))
        c2.metric("Net additional annual value", gbp_compact(hw["Net additional annual health value"]))
        st.write(f"Deadweight {pct(hw['Deadweight'])} - Displacement {pct(hw['Displacement'])} - Leakage {pct(hw['Leakage'])}")

    with st.expander("6. Civic Engagement"):
        ce = civic_engagement(place_scenario_controls, development_mix, additionality_questions, assumptions)
        sheets["6. Civic Engagement"] = pd.DataFrame(
            list({"Gross annual wellbeing value": ce["Gross annual wellbeing value"],
                  "Net additional annual wellbeing value": ce["Net additional annual wellbeing value"]}.items()),
            columns=["Measure", "Value"]).set_index("Measure")
        c1, c2, c3 = st.columns(3)
        c1.metric("Total residents", f"{ce['Total residents']:,.0f}")
        c2.metric("Deprivation uplift", f"{ce['Deprivation uplift']:.4f}")
        c3.metric("Social infrastructure uplift", f"{ce['Social infrastructure uplift']:.4f}")
        c1, c2 = st.columns(2)
        c1.metric("Gross annual value", gbp_compact(ce["Gross annual wellbeing value"]))
        c2.metric("Net additional annual value", gbp_compact(ce["Net additional annual wellbeing value"]))
        st.write(f"Deadweight {pct(ce['Deadweight'])} - Displacement {pct(ce['Displacement'])} - Leakage {pct(ce['Leakage'])}")

    with st.expander("7. Travel Time & Costs"):
        tt = travel_time_and_costs(place_scenario_controls, place_scenario_user_inputs, development_mix, additionality_questions, assumptions)
        sheets["7. Travel Time and Costs"] = pd.DataFrame(
            list({"Gross annual value": tt["Gross annual value"],
                  "Net additional annual value": tt["Net additional annual value"]}.items()),
            columns=["Measure", "Value"]).set_index("Measure")
        c1, c2, c3 = st.columns(3)
        c1.metric("Total residents (weighted)", f"{tt['Total residents (weighted)']:,.0f}")
        c2.metric("Total households", f"{tt['Total households']:,.0f}")
        c3.metric("Households with 1+ car", f"{tt['Households with 1+ car']:,.0f}")
        c1, c2, c3 = st.columns(3)
        c1.metric("Travel cost savings", gbp_compact(tt["Gross travel cost savings"]))
        c2.metric("Time savings", gbp_compact(tt["Gross time savings"]))
        c3.metric("Car ownership savings", gbp_compact(tt["Gross reduced car ownership savings"]))
        c1, c2 = st.columns(2)
        c1.metric("Gross annual value", gbp_compact(tt["Gross annual value"]))
        c2.metric("Net additional annual value", gbp_compact(tt["Net additional annual value"]))
        st.write(f"Deadweight {pct(tt['Deadweight'])} - Displacement {pct(tt['Displacement'])} - Leakage {pct(tt['Leakage'])}")

    with st.expander("8. Cost of Crime"):
        crime = cost_of_crime(crime_inputs, additionality_questions, assumptions)
        rows = []
        key_rows = []
        for category in ["Violence", "Theft", "Criminal damage", "ASB"]:
            r = crime[category]
            rows.append({"Category": category, "Incidents used": f"{r['Incidents used']:,.0f}",
                         "Reduction rate": pct(r["Reduction rate"]), "Cost/incident": gbp(r["Cost per incident"]),
                         "Gross value": gbp_compact(r["Gross value"]), "Net value": gbp_compact(r["Net value"])})
            key_rows.append({"Category": category, "Gross value": r["Gross value"], "Net value": r["Net value"]})
        crime_df = pd.DataFrame(rows).set_index("Category")
        sheets["8. Cost of Crime"] = pd.DataFrame(key_rows).set_index("Category")
        render_brand_table(crime_df.reset_index())
        c1, c2 = st.columns(2)
        c1.metric("Total gross value", gbp_compact(crime["Total"]["Gross value"]))
        c2.metric("Total net value", gbp_compact(crime["Total"]["Net value"]))
        st.write(f"Net factor: {pct(crime['Net factor'])}")

    with st.expander("9. Economic Activity"):
        econ = economic_activity(place_scenario_controls, development_mix, additionality_questions, assumptions)
        rows = []
        for row_name in ["Private detached", "Private semi/terrace", "Private low-rise flat", "Private higher density flat", "Private older persons",
                          "Social detached", "Social semi/terrace", "Social low-rise flat", "Social higher density flat", "Social older persons"]:
            r = econ[row_name]
            rows.append({"Row": row_name, "Households": f"{r['Households']:,.0f}", "Gross spend": gbp_compact(r["Gross TC spend"]),
                         "Gross jobs": f"{r['Gross FTE jobs']:,.0f}", "Gross GVA": gbp_compact(r["Gross GVA"])})
        econ_df = pd.DataFrame(rows).set_index("Row")
        render_brand_table(econ_df.reset_index())
        t = econ["Total"]
        sheets["9. Economic Activity"] = pd.DataFrame(
            list({"Gross TC spend": t["Gross TC spend"], "Gross FTE jobs": t["Gross FTE jobs"], "Gross GVA": t["Gross GVA"],
                  "Net TC spend": t["Net TC spend"], "Net FTE jobs": t["Net FTE jobs"], "Net GVA": t["Net GVA"]}.items()),
            columns=["Measure", "Value"]).set_index("Measure")
        c1, c2, c3 = st.columns(3)
        c1.metric("Total gross spend", gbp_compact(t["Gross TC spend"]))
        c2.metric("Total gross jobs", f"{t['Gross FTE jobs']:,.0f}")
        c3.metric("Total gross GVA", gbp_compact(t["Gross GVA"]))
        c1, c2, c3 = st.columns(3)
        c1.metric("Total net spend", gbp_compact(t["Net TC spend"]))
        c2.metric("Total net jobs", f"{t['Net FTE jobs']:,.0f}")
        c3.metric("Total net GVA", gbp_compact(t["Net GVA"]))
        st.write(f"Net factor: {pct(econ['Net factor'])}")

    with st.expander("10. Construction"):
        con = construction_activity(place_scenario_user_inputs, additionality_questions, assumptions)
        sheets["10. Construction"] = pd.DataFrame(
            list({"Gross PYE jobs": con["Gross PYE jobs"], "Gross GVA": con["Gross GVA"],
                  "Net additional PYE jobs": con["Net additional PYE jobs"], "Net additional GVA": con["Net additional GVA"]}.items()),
            columns=["Measure", "Value"]).set_index("Measure")
        c1, c2, c3 = st.columns(3)
        c1.metric("Capital cost", gbp_compact(con["Capital cost"]))
        c2.metric("Gross PYE jobs", f"{con['Gross PYE jobs']:,.0f}")
        c3.metric("Gross GVA", gbp_compact(con["Gross GVA"]))
        c1, c2 = st.columns(2)
        c1.metric("Net additional PYE jobs", f"{con['Net additional PYE jobs']:,.0f}")
        c2.metric("Net additional GVA", gbp_compact(con["Net additional GVA"]))
        st.write(f"Deadweight {pct(con['Deadweight'])} - Displacement {pct(con['Displacement'])} - "
                 f"Leakage {pct(con['Leakage'])} - Multiplier {con['Multiplier']:.2f}x")

    with st.expander("11. Fiscal"):
        fis = fiscal(development_mix, additionality_questions, assumptions)
        sheets["11. Fiscal"] = pd.DataFrame(
            list({"Gross fiscal value": fis["Gross fiscal value"],
                  "Net additional fiscal value": fis["Net additional fiscal value"]}.items()),
            columns=["Measure", "Value"]).set_index("Measure")
        c1, c2, c3 = st.columns(3)
        c1.metric("Gross council tax", gbp_compact(fis["Gross council tax"]))
        c2.metric("Gross rental returns", gbp_compact(fis["Gross rental returns"]))
        c3.metric("Gross fiscal value", gbp_compact(fis["Gross fiscal value"]))
        st.metric("Net additional fiscal value", gbp_compact(fis["Net additional fiscal value"]))
        st.write(f"Deadweight {pct(fis['Deadweight'])} - Displacement {pct(fis['Displacement'])} - Leakage {pct(fis['Leakage'])}")

    with st.expander("12. Land Values"):
        lv = land_values(land_infra_inputs, place_scenario_user_inputs, additionality_questions, assumptions)
        sheets["12. Land Values"] = pd.DataFrame(
            list({"Gross land value uplift": lv["Gross land value uplift"],
                  "Net additional land value uplift": lv["Net additional land value uplift"]}.items()),
            columns=["Measure", "Value"]).set_index("Measure")
        if lv["Total GDV"] == "N/A":
            st.warning("Land Values not assessed, qualifier answered 'No' in Land & Infrastructure inputs.")
        elif isinstance(lv["Gross land value uplift"], str):
            st.warning("Insufficient data, provide GDV and baseline land value in Land & Infrastructure inputs.")
        else:
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Total GDV", gbp_compact(lv["Total GDV"]))
            c2.metric("Capital cost", gbp_compact(lv["Capital cost"]))
            c3.metric("Developer return", gbp_compact(lv["Developer return"]))
            c4.metric("Baseline land value", gbp_compact(lv["Existing use / baseline land value"]))
            c1, c2 = st.columns(2)
            c1.metric("Gross land value uplift", gbp_compact(lv["Gross land value uplift"]))
            c2.metric("Net additional land value uplift", gbp_compact(lv["Net additional land value uplift"]))
            st.write(f"Deadweight {pct(lv['Deadweight'])} - Displacement {pct(lv['Displacement'])} - Leakage {pct(lv['Leakage'])}")

    with st.expander("13. Public Infrastructure"):
        pi = public_infrastructure(development_mix, assumptions)
        sheets["13. Public Infrastructure"] = pd.DataFrame(
            list({"Gross infrastructure savings (one-off)": pi["Gross infrastructure savings"],
                  "Net infrastructure savings (one-off)": pi["Net infrastructure savings"],
                  "Gross annual revenue saving": pi["Gross annual revenue saving"],
                  "Net annual revenue saving": pi["Net annual revenue saving"]}.items()),
            columns=["Measure", "Value"]).set_index("Measure")
        c1, c2 = st.columns(2)
        c1.metric("Avoided capital cost per unit", gbp_compact(pi["Avoided capital cost per unit"]))
        c2.metric("Total residential units", f"{pi['Total residential units']:,.0f}")
        c1, c2 = st.columns(2)
        c1.metric("Gross infrastructure savings (one-off)", gbp_compact(pi["Gross infrastructure savings"]))
        c2.metric("Net infrastructure savings (one-off)", gbp_compact(pi["Net infrastructure savings"]))
        c1, c2 = st.columns(2)
        c1.metric("Gross annual revenue saving", gbp_compact(pi["Gross annual revenue saving"]))
        c2.metric("Net annual revenue saving", gbp_compact(pi["Net annual revenue saving"]))
        st.write(f"Deadweight {pct(pi['Deadweight'])} - Displacement {pct(pi['Displacement'])} - Leakage {pct(pi['Leakage'])}")

    with st.expander("14. Commercial Floorspace (Optional)"):
        cf = commercial_floorspace_indicator(commercial_floorspace_inputs, additionality_questions, assumptions)
        rows = []
        key_rows = []
        for category in commercial_floorspace_inputs:
            r = cf[category]
            rows.append({"Category": category, "Floorspace (m²)": f"{r['Floorspace']:,.0f}",
                         "FTE jobs": f"{r['FTE jobs']:,.0f}", "Gross GVA": gbp_compact(r["Gross GVA"])})
            key_rows.append({"Category": category, "FTE jobs": r["FTE jobs"], "Gross GVA": r["Gross GVA"]})
        cf_df = pd.DataFrame(rows).set_index("Category")
        sheets["14. Commercial Floorspace"] = pd.DataFrame(key_rows).set_index("Category")
        render_brand_table(cf_df.reset_index())
        t = cf["Total"]
        o = cf["Occupancy-adjusted (75%)"]
        c1, c2 = st.columns(2)
        c1.metric("Total gross jobs", f"{t['Gross FTE jobs']:,.0f}")
        c2.metric("Total gross GVA", gbp_compact(t["Gross GVA"]))
        c1, c2 = st.columns(2)
        c1.metric("Occupancy-adjusted (75%) jobs", f"{o['Gross FTE jobs']:,.0f}")
        c2.metric("Occupancy-adjusted (75%) GVA", gbp_compact(o["Gross GVA"]))

    with st.expander("15. Purpose Built Student Accommodation"):
        pbsa = pbsa_indicator(pbsa_inputs, additionality_questions, assumptions)
        sheets["15. PBSA"] = pd.DataFrame(
            list({"Gross total FTE jobs": pbsa["Gross total FTE jobs"], "Gross total GVA": pbsa["Gross total GVA"],
                  "Net total FTE jobs": pbsa["Net total FTE jobs"], "Net total GVA": pbsa["Net total GVA"],
                  "Net total spend": pbsa["Net total spend"]}.items()),
            columns=["Measure", "Value"]).set_index("Measure")
        c1, c2 = st.columns(2)
        c1.metric("Number of rooms", f"{pbsa['Number of rooms']:,.0f}")
        c2.metric("On-site FTE jobs (e.g. concierge)", f"{pbsa['On-site FTE jobs (e.g. concierge)']:,.0f}")
        c1, c2 = st.columns(2)
        c1.metric("Total off-site spend", gbp_compact(pbsa["Total off-site spend"]))
        c2.metric("Off-site FTE jobs", f"{pbsa['Off-site FTE jobs']:,.0f}")
        c1, c2 = st.columns(2)
        c1.metric("Gross total FTE jobs", f"{pbsa['Gross total FTE jobs']:,.0f}")
        c2.metric("Gross total GVA", gbp_compact(pbsa["Gross total GVA"]))
        c1, c2 = st.columns(2)
        c1.metric("Net total FTE jobs", f"{pbsa['Net total FTE jobs']:,.0f}")
        c2.metric("Net total GVA", gbp_compact(pbsa["Net total GVA"]))
        st.write(f"Deadweight {pct(pbsa['Deadweight'])} - Displacement {pct(pbsa['Displacement'])} - Multiplier {pbsa['Multiplier']:.2f}x")


    st.divider()
    st.download_button(
        "Download full report (.xlsx)",
        data=build_excel_export(dashboard, sheets),
        file_name="tcl_impact_report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
        key="download_final",
    )